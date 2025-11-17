from django.shortcuts import render, redirect, get_object_or_404
from .forms import *
from .models import *
from django.db.models import Q, Sum, F, Value, CharField, ExpressionWrapper, DecimalField
from django.db.models.functions import TruncYear, TruncQuarter, TruncMonth, TruncDay, Coalesce, Cast
import csv
from datetime import datetime
from django.http import HttpResponse
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.contrib import messages
from django.contrib.auth.decorators import login_required, permission_required
from django.http import JsonResponse
import json
from decimal import Decimal
from django.forms import modelformset_factory
from django.db.models import Sum, Avg, Max, Min, Count
from django.db.models.functions import TruncMonth
from datetime import datetime
from collections import defaultdict
from django.contrib.auth import login, logout
from django.contrib.auth.views import LoginView
from django.urls import reverse_lazy

# Create your views here.
#Logic for main page
@login_required
def main(request):
    # Obtener cheques y echeques próximos a vencer (próximos 30 días)
    from datetime import datetime, timedelta
    
    fecha_limite = datetime.now().date() + timedelta(days=30)
    
    # Obtener cheques de ingresos pendientes
    cheques_pendientes = IngresoFinanciero.objects.filter(
        forma_pago__in=['Cheque', 'Echeque'],
        fecha_pago__isnull=False,
        fecha_pago__gte=datetime.now().date(),  # Solo fechas futuras o de hoy
        fecha_pago__lte=fecha_limite  # Dentro de los próximos 30 días
    ).select_related().order_by('fecha_pago')
    
    context = {
        'cheques_pendientes': cheques_pendientes,
        'fecha_limite': fecha_limite,
        'total_cheques': cheques_pendientes.count(),
        'monto_total_pendiente': sum(cheque.monto for cheque in cheques_pendientes)
    }
    
    return render(request, 'contabilidad_loslirios/main.html', context)

#API endpoint to get all parcelas in GeoJSON format
def parcelas_geojson(request):
    """
    Esta vista devuelve todas las parcelas en formato GeoJSON.
    """
    parcelas = Parcela.objects.all()
    
    # Creamos la estructura base de un FeatureCollection de GeoJSON
    features = []
    for parcela in parcelas:
        if parcela.coordenadas: # Solo incluimos parcelas con coordenadas
            features.append({
                "type": "Feature",
                "geometry": {
                    "type": "Polygon",
                    # El formato GeoJSON para polígonos requiere una lista de anillos, 
                    # el primero es el contorno exterior.
                    "coordinates": [
                        # Invertimos [lat, lon] a [lon, lat] como lo espera GeoJSON
                        [[lon, lat] for lat, lon in parcela.coordenadas]
                    ]
                },
                "properties": {
                    "nombre": parcela.nombre,
                    "variedad": parcela.variedad or "N/D",
                    "superficie_ha": parcela.superficie_ha or "N/D",
                    "cabezal_riego": parcela.cabezal_riego or "N/D",
                }
            })
            
    geojson_data = {
        "type": "FeatureCollection",
        "features": features
    }
    
    return JsonResponse(geojson_data)

#Logic for login/logout
class CustomLoginView(LoginView):
    form_class = CustomLoginForm
    template_name = 'registration/login.html'
    success_url = reverse_lazy('main')
    
    def get_success_url(self):
        return reverse_lazy('main')
    
    def form_valid(self, form):
        """Agregar mensaje de bienvenida"""
        response = super().form_valid(form)
        messages.success(self.request, f'¡Bienvenido de nuevo, {self.request.user.username}!')
        return response

@login_required
def custom_logout(request):
    username = request.user.username
    logout(request)
    messages.success(request, f'Sesión cerrada exitosamente. ¡Hasta luego, {username}!')
    return redirect('login')


#Logic for administracion page:
@permission_required(['contabilidad_loslirios.can_view_jornales', 'contabilidad_loslirios.can_view_movimientos'], raise_exception=True)
@login_required
def contabilidad(request):
    return render(request, 'contabilidad_loslirios/contabilidad.html')

#Logic for movimientos
# API endpoint to get classifications based on type
def get_classifications_for_type(request, tipo):
    classifications = CLASIFICACIONES_POR_TIPO.get(tipo, [])
    return JsonResponse({'classifications': classifications})

#Logic for cargar_movimiento page
@permission_required('contabilidad_loslirios.can_add_movimientos', raise_exception=True)
@login_required
def cargar_movimiento(request):
    if request.method == 'POST':
        formulario = FormMovimientoFinanciero(request.POST)
        if formulario.is_valid():
            formulario.save()
            messages.success(request, 'Movimiento financiero guardado exitosamente.')
            return redirect('cargar_movimiento') 
        else:
            messages.error(request, f'Error al guardar el movimiento financiero: {formulario.errors.as_text()}')
    else: 
        formulario = FormMovimientoFinanciero() 
    
    return render(request, 'contabilidad_loslirios/administracion/cargar_movimiento.html', {'formulario': formulario})

#Logic for consultar_movimiento page:
@permission_required('contabilidad_loslirios.can_view_movimientos', raise_exception=True)
@login_required
def consultar_movimiento(request):
    movimientos_filtrados = _obtener_movimientos_filtrados(request)
    form = FormConsultaMovimiento(request.GET)

    # --- Lógica de Paginación ---
    paginator = Paginator(movimientos_filtrados, 4)  
    page = request.GET.get('page')

    try:
        movimientos_paginados = paginator.page(page)
    except PageNotAnInteger:
        movimientos_paginados = paginator.page(1)
    except EmptyPage:
        movimientos_paginados = paginator.page(paginator.num_pages)

    context = {
        'form': form,
        'movimientos': movimientos_paginados, 
        'paginator': paginator, 
    }
    return render(request, 'contabilidad_loslirios/administracion/consultar_movimiento.html', context)

#Logic for exportar_movimiento page
#Auxiliary function to export filtered records to CSV
@login_required
def _obtener_movimientos_filtrados(request):
    form = FormConsultaMovimiento(request.GET)
    movimientos = MovimientoFinanciero.objects.all() # El ordenado ya está en el Meta del modelo

    if form.is_valid():
        filtros = Q()
        # Añadimos todos los nuevos filtros
        for field, value in form.cleaned_data.items():
            if value:
                if field in ['fecha_desde']:
                    filtros &= Q(fecha__gte=value)
                elif field in ['fecha_hasta']:
                    filtros &= Q(fecha__lte=value)
                else:
                    filtros &= Q(**{f'{field}__exact': value})

        movimientos = movimientos.filter(filtros)

    return movimientos

#import csv
@permission_required('contabilidad_loslirios.can_export_movimientos', raise_exception=True)
@login_required
def exportar_movimientos_csv(request):
    movimientos = _obtener_movimientos_filtrados(request)
    response = HttpResponse(content_type='text/csv')
    filename = f"movimientos_consulta_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    writer = csv.writer(response)

    # Nuevos encabezados
    writer.writerow(['Fecha', 'Origen', 'Finca', 'Tipo', 'Clasificacion', 'Detalle', 'Monto', 'Moneda', 'Forma de Pago'])

    # Nuevos datos
    for m in movimientos:
        writer.writerow([
            m.fecha.strftime('%Y-%m-%d'), m.origen, m.finca, m.tipo, m.clasificacion, m.detalle, f"{m.monto:.2f}", m.moneda, m.forma_pago
        ])
    return response

#Logic for ingresos page:
def get_destinos_ingresos(request):
    """API para obtener todos los destinos disponibles"""
    # Destinos predefinidos
    destinos_predefinidos = [choice[0] for choice in DESTINO_INGRESOS_CHOICES]
    
    # Destinos personalizados
    destinos_personalizados = list(DestinoIngreso.objects.values_list('nombre', flat=True))
    
    # Combinar y eliminar duplicados
    todos_destinos = list(set(destinos_predefinidos + destinos_personalizados))
    todos_destinos.sort()
    
    return JsonResponse({'destinos': todos_destinos})

def get_compradores_ingresos(request):
    """API para obtener todos los compradores disponibles"""
    # Compradores personalizados (por ahora no hay predefinidos)
    compradores = list(CompradorIngreso.objects.values_list('nombre', flat=True))
    compradores.sort()
    
    return JsonResponse({'compradores': compradores})

def agregar_destino_ingreso(request):
    """API para agregar un nuevo destino personalizado"""
    if request.method == 'POST':
        nombre = request.POST.get('nombre', '').strip()
        if nombre:
            destino, created = DestinoIngreso.objects.get_or_create(nombre=nombre)
            if created:
                return JsonResponse({'success': True, 'message': 'Destino agregado exitosamente'})
            else:
                return JsonResponse({'success': False, 'message': 'El destino ya existe'})
    return JsonResponse({'success': False, 'message': 'Nombre requerido'})

def agregar_comprador_ingreso(request):
    """API para agregar un nuevo comprador personalizado"""
    if request.method == 'POST':
        nombre = request.POST.get('nombre', '').strip()
        if nombre:
            comprador, created = CompradorIngreso.objects.get_or_create(nombre=nombre)
            if created:
                return JsonResponse({'success': True, 'message': 'Comprador agregado exitosamente'})
            else:
                return JsonResponse({'success': False, 'message': 'El comprador ya existe'})
    return JsonResponse({'success': False, 'message': 'Nombre requerido'})

#Logic for cargar_ingresos page:
@permission_required('contabilidad_loslirios.can_view_ingresos', raise_exception=True)
@login_required
def cargar_ingresos(request):
    if request.method == 'POST':
        form = FormIngresoFinanciero(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Ingreso financiero guardado exitosamente.')
            return redirect('cargar_ingresos')
        else:
            messages.error(request, f'Error al guardar el ingreso financiero: {form.errors.as_text()}')
    else:
        form = FormIngresoFinanciero()

    return render(request, 'contabilidad_loslirios/administracion/cargar_ingresos.html', {'formulario': form})
#Logic for consultar_ingresos page:
@permission_required('contabilidad_loslirios.can_view_ingresos', raise_exception=True)
@login_required
def consultar_ingresos(request):
    ingresos_filtrados = _obtener_ingresos_filtrados(request)
    form = FormConsultaIngresos(request.GET)  

    # --- Lógica de Paginación ---
    paginator = Paginator(ingresos_filtrados, 4)  
    page = request.GET.get('page')

    try:
        ingresos_paginados = paginator.page(page)
    except PageNotAnInteger:
        ingresos_paginados = paginator.page(1)
    except EmptyPage:
        ingresos_paginados = paginator.page(paginator.num_pages)

    context = {
        'form': form,
        'ingresos': ingresos_paginados, 
        'paginator': paginator, 
    }
    return render(request, 'contabilidad_loslirios/administracion/consultar_ingresos.html', context)

#Logic for exportar_ingresos page
#Auxiliary function to export filtered ingresos to CSV
@login_required
def _obtener_ingresos_filtrados(request):
    form = FormConsultaIngresos(request.GET)  # Cambiar nombre del formulario
    ingresos = IngresoFinanciero.objects.all() 

    if form.is_valid():
        filtros = Q()
        for field, value in form.cleaned_data.items():
            if value:
                if field == 'fecha_desde':
                    filtros &= Q(fecha__gte=value)
                elif field == 'fecha_hasta':
                    filtros &= Q(fecha__lte=value)
                elif field in ['destino', 'comprador']:
                    # Búsqueda parcial para campos de texto
                    filtros &= Q(**{f'{field}__icontains': value})
                else:
                    # Búsqueda exacta para campos de selección
                    filtros &= Q(**{f'{field}__exact': value})

        ingresos = ingresos.filter(filtros)

    return ingresos

#import csv
@permission_required('contabilidad_loslirios.can_export_ingresos', raise_exception=True)
@login_required
def exportar_ingresos_csv(request):
    ingresos = _obtener_ingresos_filtrados(request)
    response = HttpResponse(content_type='text/csv')
    filename = f"ingresos_consulta_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    writer = csv.writer(response)

    # NUEVOS encabezados actualizados:
    writer.writerow([
        'Fecha', 'Origen', 'Destino', 'Comprador', 'Forma de Pago', 'Forma', 
        'Banco', 'N° de Cheque', 'Fecha de Pago', 'Monto', 'Moneda'
    ])

    # NUEVOS datos actualizados:
    for ingreso in ingresos:
        writer.writerow([
            ingreso.fecha.strftime('%Y-%m-%d'), 
            ingreso.origen, 
            ingreso.destino, 
            ingreso.comprador,
            ingreso.forma_pago,
            ingreso.forma,
            ingreso.banco or '',
            ingreso.numero_cheque or '',
            ingreso.fecha_pago.strftime('%Y-%m-%d') if ingreso.fecha_pago else '',
            f"{ingreso.monto:.2f}", 
            ingreso.moneda
        ])
    return response




#Logic for produccion page:
@permission_required('contabilidad_loslirios.can_view_produccion_data', raise_exception=True) 
@login_required
def produccion(request):
    return render(request, 'contabilidad_loslirios/produccion.html')

#Logic for Jornales
#API endpoint to get tasks based on classification
def get_tasks_for_classification(request, classification):
    """
    Devuelve una lista de tareas en formato JSON para una clasificación dada.
    """
    tasks = TAREAS_POR_CLASIFICACION.get(classification, [])
    return JsonResponse({'tasks': tasks})

#Logic for cargar_jornal page:
@permission_required('contabilidad_loslirios.can_add_jornales', raise_exception=True)
@login_required
def cargar_jornal(request):
    RegistroTrabajoFormSet = modelformset_factory(
        registro_trabajo,
        form=FormRegistroTrabajo,
        extra=3,  
        can_delete=False
    )
    if request.method == 'POST':
        formset = RegistroTrabajoFormSet(request.POST)
        if formset.is_valid():
            formset.save()
            messages.success(request, 'Registros de jornal guardados exitosamente.')
            return redirect('cargar_jornal')
        else:
            messages.error(request, 'Error al guardar los registros de jornal.')
    else:
        formset = RegistroTrabajoFormSet(queryset=registro_trabajo.objects.none())
    return render(request, 'contabilidad_loslirios/produccion/cargar_jornal.html', {'formset': formset})

#Logic for consultar_jornal page:
@permission_required('contabilidad_loslirios.can_view_jornales', raise_exception=True)
@login_required
def consultar_jornal(request):
    registros_filtrados = _obtener_registros_filtrados(request)
    form = FormConsultaJornal(request.GET)

    # --- Lógica de Paginación ---
    paginator = Paginator(registros_filtrados, 4)
    page = request.GET.get('page') 

    try:
        registros_paginados = paginator.page(page)
    except PageNotAnInteger:
        registros_paginados = paginator.page(1)
    except EmptyPage:
        registros_paginados = paginator.page(paginator.num_pages)

    context = {
        'form': form,
        'registros': registros_paginados,
        'paginator': paginator,
    }
    return render(request, 'contabilidad_loslirios/produccion/consultar_jornal.html', context)

# Logic for exportar_jornal page
#Auxiliary function to export filtered records to CSV
@login_required
def _obtener_registros_filtrados(request):
    """
    Función auxiliar para obtener los registros filtrados basada en los parámetros GET.
    Reutiliza la lógica de filtrado de consultar_jornal.
    """
    form = FormConsultaJornal(request.GET)
    registros = registro_trabajo.objects.all().order_by('-fecha')

    if form.is_valid():
        fecha_desde = form.cleaned_data.get('fecha_desde')
        fecha_hasta = form.cleaned_data.get('fecha_hasta')
        nombre_trabajador = form.cleaned_data.get('nombre_trabajador')
        tarea = form.cleaned_data.get('tarea')
        ubicacion = form.cleaned_data.get('ubicacion')
        clasificacion = form.cleaned_data.get('clasificacion')
        detalle = form.cleaned_data.get('detalle') 
        monto_total = form.cleaned_data.get('monto_total')

        filtros = Q()

        if fecha_desde:
            filtros &= Q(fecha__gte=fecha_desde)
        if fecha_hasta:
            filtros &= Q(fecha__lte=fecha_hasta)
        if nombre_trabajador:
            filtros &= Q(nombre_trabajador__icontains=nombre_trabajador)
        if tarea:
            filtros &= Q(tarea=tarea) 
        if ubicacion:
            filtros &= Q(ubicacion__icontains=ubicacion) 
        if clasificacion:
            filtros &= Q(clasificacion=clasificacion)
        if detalle:
            filtros &= Q(detalle__icontains=detalle) 

        registros = registros.filter(filtros)

    return registros

#import csv
@permission_required('contabilidad_loslirios.can_export_jornales', raise_exception=True)
@login_required
def exportar_jornales_csv(request):
    registros = _obtener_registros_filtrados(request)

    response = HttpResponse(content_type='text/csv')
    filename = f"jornales_consulta_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    writer = csv.writer(response)

    # Encabezados del CSV 
    writer.writerow(['Fecha', 'Nombre Trabajador', 'Clasificacion', 'Tarea', 'Detalle', 'Cantidad', 'Unidad Medida', 'Precio', 'Ubicacion'])

    # Datos
    for registro in registros:
        writer.writerow([
            registro.fecha.strftime('%Y-%m-%d'),
            registro.nombre_trabajador,
            registro.clasificacion,
            registro.tarea,
            registro.detalle, 
            registro.cantidad,
            registro.unidad_medida,
            f"{registro.precio:.2f}", 
            registro.ubicacion,
            registro.monto_total,
        ])
    return response


#Logic for riego y fertilización
# API endpoint to get parrales based on the selected cabezal
def get_parrales_for_cabezal(request, cabezal):
    """
    Devuelve una lista de parrales en formato JSON para un cabezal dado.
    """
    parrales = list(RIEGO_DATA.get(cabezal, {}).keys())
    return JsonResponse({'parrales': parrales})

# API endpoint to get valves based on the selected cabezal and parral
def get_valvulas_for_parral(request, cabezal, parral):
    """
    Devuelve una lista de válvulas en formato JSON para un parral y cabezal dados.
    """
    valvulas = RIEGO_DATA.get(cabezal, {}).get(parral, [])
    return JsonResponse({'valvulas': valvulas})

# Logic for cargar_riego page
@permission_required('contabilidad_loslirios.can_add_riego', raise_exception=True)
@login_required
def cargar_riego(request):
    if request.method == 'POST':
        form = FormRegistroRiego(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Registro de riego guardado exitosamente.')
            return redirect('cargar_riego')
        else:
            # Capturamos los errores para mostrarlos
            error_text = form.errors.as_text()
            messages.error(request, f'Error al guardar el registro: {error_text}')
    else:
        form = FormRegistroRiego()
    
    context = {
        'form': form
    }
    return render(request, 'contabilidad_loslirios/produccion/cargar_riego.html', context)

# Logic for consultar_riego page
@permission_required('contabilidad_loslirios.can_view_riego', raise_exception=True)
@login_required
def consultar_riego(request):
    # Crear formulario de filtros
    form_filtros = FiltrosRiegoForm(request.GET or None)
    
    # Poblar opciones dinámicas
    cabezales = RegistroRiego.objects.values_list('cabezal', flat=True).distinct().order_by('cabezal')
    form_filtros.fields['cabezal'].choices = [('', 'Todos')] + [(c, c) for c in cabezales if c]
    
    # Obtener registros base
    registros = RegistroRiego.objects.all()
    
    # Aplicar filtros
    if form_filtros.is_valid():
        fecha_desde = form_filtros.cleaned_data.get('fecha_desde')
        fecha_hasta = form_filtros.cleaned_data.get('fecha_hasta')
        cabezal = form_filtros.cleaned_data.get('cabezal')
        parral = form_filtros.cleaned_data.get('parral')
        responsable = form_filtros.cleaned_data.get('responsable')
        
        if fecha_desde:
            registros = registros.filter(inicio__date__gte=fecha_desde)
        if fecha_hasta:
            registros = registros.filter(inicio__date__lte=fecha_hasta)
        if cabezal:
            registros = registros.filter(cabezal=cabezal)
        if parral:
            registros = registros.filter(parral__icontains=parral)
        if responsable:
            registros = registros.filter(responsable__icontains=responsable)
    
    # Paginación
    paginator = Paginator(registros.order_by('-inicio'), 5)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'form_filtros': form_filtros,  # Cambiar nombre del formulario
        'page_obj': page_obj,
    }
    return render(request, 'contabilidad_loslirios/produccion/consultar_riego.html', context)

# Logic for exportar_riego page 
@permission_required('contabilidad_loslirios.can_view_riego', raise_exception=True) # O un nuevo permiso de exportación si lo creas
@login_required
def exportar_riegos_csv(request):
    # Crear la respuesta HTTP con tipo de contenido CSV
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename="registros_riego.csv"'
    
    # Configurar el writer CSV con codificación UTF-8
    response.write('\ufeff'.encode('utf8'))  # BOM para Excel
    writer = csv.writer(response)
    
    # Escribir encabezados 
    writer.writerow([
        'Fecha/Hora Inicio', 
        'Fecha/Hora Fin', 
        'Cabezal', 
        'Parral/Potrero', 
        'Válvula Abierta',
        'Total Horas',
        'Fertilizante',
        'Litros Fertilizante',
        'Responsable'
    ])
    
    # Obtener los registros con los mismos filtros que la vista de consulta
    form_filtros = FiltrosRiegoForm(request.GET or None)
    registros = RegistroRiego.objects.all()
    
    # Aplicar filtros si el formulario es válido
    if form_filtros.is_valid():
        fecha_desde = form_filtros.cleaned_data.get('fecha_desde')
        fecha_hasta = form_filtros.cleaned_data.get('fecha_hasta')
        cabezal = form_filtros.cleaned_data.get('cabezal')
        parral = form_filtros.cleaned_data.get('parral')
        responsable = form_filtros.cleaned_data.get('responsable')
        
        if fecha_desde:
            registros = registros.filter(inicio__date__gte=fecha_desde)
        if fecha_hasta:
            registros = registros.filter(inicio__date__lte=fecha_hasta)
        if cabezal:
            registros = registros.filter(cabezal=cabezal)
        if parral:
            registros = registros.filter(parral__icontains=parral)
        if responsable:
            registros = registros.filter(responsable__icontains=responsable)
    
    # Escribir los datos 
    for r in registros.order_by('-inicio'):
        writer.writerow([
            r.inicio.strftime('%d/%m/%Y %H:%M') if r.inicio else '',
            r.fin.strftime('%d/%m/%Y %H:%M') if r.fin else '',
            r.cabezal or '',
            r.parral or '',
            r.valvula_abierta or '',
            r.total_horas or '',
            r.fertilizante_nombre or '',
            r.fertilizante_litros or '',
            r.responsable or ''
            # Removido r.observaciones
        ])  
    return response
# Auxiliary function to get filtered irrigation records
@login_required
def _obtener_riegos_filtrados(request):
    """
    Función auxiliar para obtener los registros de riego filtrados.
    """
    registros = RegistroRiego.objects.all()
    form = FiltrosRiegoForm(request.GET)

    if form.is_valid():
        filtros = Q()
        if form.cleaned_data.get('fecha_desde'):
            # Filtramos por la parte de la fecha del campo DateTimeField
            filtros &= Q(inicio__date__gte=form.cleaned_data['fecha_desde'])
        if form.cleaned_data.get('fecha_hasta'):
            filtros &= Q(inicio__date__lte=form.cleaned_data['fecha_hasta'])
        if form.cleaned_data.get('cabezal'):
            filtros &= Q(cabezal__exact=form.cleaned_data['cabezal'])
        if form.cleaned_data.get('parral'):
            filtros &= Q(parral__exact=form.cleaned_data['parral'])
        if form.cleaned_data.get('responsable'):
            filtros &= Q(responsable__icontains=form.cleaned_data['responsable'])
        
        registros = registros.filter(filtros)
        
    return registros

#Logic for cosecha page:
#Logic for cargar_cosecha page:
@login_required
def cargar_cosecha(request):
    if request.method == 'POST':
        form = RegistroCosechaForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Registro de cosecha guardado exitosamente.')
            return redirect('cargar_cosecha')
        else:
            messages.error(request, 'Error al guardar el registro. Por favor, revise los datos.')
    else:
        form = RegistroCosechaForm()
    
    # Obtener opciones para datalists
    fincas = list(set([choice[0] for choice in FINCA_CHOICES] + 
                     list(RegistroCosecha.objects.values_list('finca', flat=True).distinct())))
    compradores = list(set([choice[0] for choice in COMPRADOR_CHOICES] + 
                          list(RegistroCosecha.objects.values_list('comprador', flat=True).distinct())))
    cultivos = list(set([choice[0] for choice in CULTIVO_CHOICES] + 
                       list(RegistroCosecha.objects.values_list('cultivo', flat=True).distinct())))
    variedades = list(set([choice[0] for choice in VARIEDAD_CHOICES] + 
                         list(RegistroCosecha.objects.values_list('variedad', flat=True).distinct())))
    
    context = {
        'form': form,
        'fincas': fincas,
        'compradores': compradores,
        'cultivos': cultivos,
        'variedades': variedades,
    }
    return render(request, 'contabilidad_loslirios/produccion/cargar_cosecha.html', context)

#Logic for consultar_cosecha page:
@login_required
def consultar_cosecha(request):
    # Inicializar formulario de filtros
    form_filtros = FiltrosCosechaForm(request.GET or None)
    
    # Obtener todos los registros
    registros = RegistroCosecha.objects.all()
    
    # Aplicar filtros si el formulario es válido
    if form_filtros.is_valid():
        fecha_inicio = form_filtros.cleaned_data.get('fecha_inicio')
        fecha_fin = form_filtros.cleaned_data.get('fecha_fin')
        origen = form_filtros.cleaned_data.get('origen')
        finca = form_filtros.cleaned_data.get('finca')
        destino = form_filtros.cleaned_data.get('destino')
        comprador = form_filtros.cleaned_data.get('comprador')
        variedad = form_filtros.cleaned_data.get('variedad')
        
        if fecha_inicio:
            registros = registros.filter(fecha__gte=fecha_inicio)
        if fecha_fin:
            registros = registros.filter(fecha__lte=fecha_fin)
        if origen:
            registros = registros.filter(origen=origen)
        if finca:
            registros = registros.filter(finca__icontains=finca)
        if destino:
            registros = registros.filter(destino=destino)
        if comprador:
            registros = registros.filter(comprador__icontains=comprador)
        if variedad:
            registros = registros.filter(variedad__icontains=variedad)
    
    # Calcular totales antes de la paginación
    total_registros = registros.count()
    total_kg = registros.aggregate(Sum('kg_totales'))['kg_totales__sum'] or 0
    
    # Paginación
    paginator = Paginator(registros.order_by('-fecha', '-id'), 5)  
    page_number = request.GET.get('page')
    registros_paginados = paginator.get_page(page_number)
    
    context = {
        'form_filtros': form_filtros,
        'registros': registros_paginados,
        'total_registros': total_registros,
        'total_kg': total_kg,
        'paginator': paginator,
    }
    return render(request, 'contabilidad_loslirios/produccion/consultar_cosecha.html', context)

#Logic for exportar_cosecha page:
@login_required
def exportar_cosecha_csv(request):
    # Aplicar los mismos filtros que en consultar
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    origen = request.GET.get('origen')
    finca = request.GET.get('finca')
    destino = request.GET.get('destino')
    variedad = request.GET.get('variedad')
    
    registros = RegistroCosecha.objects.all()
    
    if fecha_inicio:
        registros = registros.filter(fecha__gte=fecha_inicio)
    if fecha_fin:
        registros = registros.filter(fecha__lte=fecha_fin)
    if origen:
        registros = registros.filter(origen=origen)
    if finca:
        registros = registros.filter(finca__icontains=finca)
    if destino:
        registros = registros.filter(destino=destino)
    if variedad:
        registros = registros.filter(variedad__icontains=variedad)
    
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="registros_cosecha.csv"'
    
    writer = csv.writer(response)
    writer.writerow([
        'ID', 'Fecha', 'Origen', 'Finca', 'Destino', 'Comprador', 'Cultivo',
        'Parral/Potrero', 'Variedad', 'Remito', 'CIU', 'Medida', 'Peso Unitario',
        'Cantidad', 'Bruto', 'Tara', 'Kg Totales'
    ])
    
    for registro in registros:
        writer.writerow([
            registro.id, registro.fecha, registro.origen, registro.finca,
            registro.destino, registro.comprador, registro.cultivo,
            registro.parral_potrero, registro.variedad, registro.remito,
            registro.ciu or '', registro.medida, registro.peso_unitario or '',
            registro.cantidad, registro.bruto or '', registro.tara or '',
            registro.kg_totales
        ])
    return response


#Logic for analisis page:
#Logic for dashbooard jornales page:
@permission_required('contabilidad_loslirios.can_view_analisis_data', raise_exception=True)
@login_required
def analisis(request):
    # --- 1. PROCESAR FILTROS ---
    form = FormFiltroDashboardJornales(request.GET or None)
    queryset = registro_trabajo.objects.all()

    if form.is_valid():
        if form.cleaned_data.get('fecha_desde'):
            queryset = queryset.filter(fecha__gte=form.cleaned_data['fecha_desde'])
        if form.cleaned_data.get('fecha_hasta'):
            queryset = queryset.filter(fecha__lte=form.cleaned_data['fecha_hasta'])
        if form.cleaned_data.get('clasificacion'):
            queryset = queryset.filter(clasificacion__in=form.cleaned_data['clasificacion'])
        if form.cleaned_data.get('tarea'):
            queryset = queryset.filter(tarea__in=form.cleaned_data['tarea'])
        if form.cleaned_data.get('ubicacion'):
            queryset = queryset.filter(ubicacion__in=form.cleaned_data['ubicacion'])

    # --- 2. CALCULAR KPIs MEJORADOS ---
    costo_total_calculado = queryset.annotate(
        costo=ExpressionWrapper(
            F('cantidad') * F('precio'),
            output_field=DecimalField()
        )
    ).aggregate(total=Coalesce(Sum('costo'), Decimal('0.00'), output_field=DecimalField()))
    
    # CORREGIDO: Trabajador más productivo
    trabajador_mas_productivo = queryset.annotate(
        costo=ExpressionWrapper(
            F('cantidad') * F('precio'),
            output_field=DecimalField()
        )
    ).values('nombre_trabajador').annotate(
        total=Sum('costo')
    ).order_by('-total').first()
    
    # Costo por hectárea (asumiendo que las ubicaciones representan hectáreas)
    ubicaciones_count = queryset.values('ubicacion').distinct().count()
    costo_por_hectarea = (costo_total_calculado['total'] / ubicaciones_count) if ubicaciones_count > 0 else 0

    kpis = {
        'costo_total': costo_total_calculado['total'],
        'total_registros': queryset.count(),
        'trabajadores_activos': queryset.values('nombre_trabajador').distinct().count(),
        'trabajador_mas_productivo': trabajador_mas_productivo,  # Ya contiene nombre_trabajador y total
        'costo_por_hectarea': costo_por_hectarea,
    }

    # --- 3. EFICIENCIA POR TRABAJADOR (TOP 5) ---
    eficiencia_trabajadores = queryset.annotate(
        costo=ExpressionWrapper(
            F('cantidad') * F('precio'),
            output_field=DecimalField()
        )
    ).values('nombre_trabajador').annotate(
        total_costo=Sum('costo'),
        registros=Count('id_registro'),
        eficiencia=Avg('costo')
    ).order_by('-total_costo')[:5]

    # --- 4. RESUMEN POR TAREA ---
    resumen_tareas = queryset.annotate(
        costo=ExpressionWrapper(
            F('cantidad') * F('precio'),
            output_field=DecimalField()
        )
    ).values('tarea').annotate(
        registros=Count('id_registro'),
        costo_total=Sum('costo'),
        costo_promedio=Avg('costo'),
        eficiencia=Avg('costo')
    ).order_by('-costo_total')

    # Calcular porcentajes y eficiencia relativa
    costo_total_global = sum([t['costo_total'] for t in resumen_tareas]) if resumen_tareas else 1
    eficiencia_max = max([t['eficiencia'] for t in resumen_tareas]) if resumen_tareas else 1

    for tarea in resumen_tareas:
        tarea['porcentaje_total'] = (tarea['costo_total'] / costo_total_global * 100) if costo_total_global > 0 else 0
        tarea['eficiencia_porcentaje'] = (tarea['eficiencia'] / eficiencia_max * 100) if eficiencia_max > 0 else 0

    # --- 5. DATOS PARA GRÁFICOS ---
    agrupacion = request.GET.get('agrupacion', 'mes')

    if agrupacion == 'anio':
        trunc_func = TruncYear('fecha')
        date_format = lambda d: d.strftime('%Y')
    elif agrupacion == 'trimestre':
        trunc_func = TruncQuarter('fecha')
        date_format = lambda d: f"T{((d.month-1)//3)+1} {d.year}"
    elif agrupacion == 'dia':
        trunc_func = TruncDay('fecha')
        date_format = lambda d: d.strftime('%d/%m/%Y')
    else:  # mes por defecto
        trunc_func = TruncMonth('fecha')
        date_format = lambda d: d.strftime('%b %Y')

    costo_mensual = queryset.annotate(
        costo=ExpressionWrapper(
            F('cantidad') * F('precio'),
            output_field=DecimalField()
        ),
        periodo=trunc_func
    ).values('periodo').annotate(total_costo=Sum('costo')).order_by('periodo')

    line_chart_labels = [date_format(c['periodo']) for c in costo_mensual]
    line_chart_data = [float(c['total_costo']) for c in costo_mensual]

    # Gráfico de Barras
    costo_por_tarea = queryset.annotate(
        costo=ExpressionWrapper(
            Cast(F('cantidad'), output_field=DecimalField()) * Cast(F('precio'), output_field=DecimalField()),
            output_field=DecimalField()
        )
    ).values('tarea').annotate(total_costo=Sum('costo')).order_by('-total_costo')[:10]  # Top 10
    bar_chart_labels = [c['tarea'] for c in costo_por_tarea]
    bar_chart_data = [float(c['total_costo']) for c in costo_por_tarea]
    
    # Gráfico de Torta
    costo_por_clasificacion = queryset.annotate(
        costo=ExpressionWrapper(
            Cast(F('cantidad'), output_field=DecimalField()) * Cast(F('precio'), output_field=DecimalField()),
            output_field=DecimalField()
        )
    ).values('clasificacion').annotate(total_costo=Sum('costo')).order_by('-total_costo')
    pie_chart_labels = [c['clasificacion'] for c in costo_por_clasificacion]
    pie_chart_data = [float(c['total_costo']) for c in costo_por_clasificacion]
    
    # Gráfico de Radar (distribución por temporada)
    radar_labels = pie_chart_labels  # Usar las mismas clasificaciones
    radar_data = pie_chart_data
    
    # --- CONTEXTO PARA LA PLANTILLA ---
    context = {
        'form': form,
        'kpis': kpis,
        'eficiencia_trabajadores': eficiencia_trabajadores,
        'resumen_tareas': resumen_tareas,
        'line_chart_labels': json.dumps(line_chart_labels),
        'line_chart_data': json.dumps(line_chart_data),
        'bar_chart_labels': json.dumps(bar_chart_labels),
        'bar_chart_data': json.dumps(bar_chart_data),
        'pie_chart_labels': json.dumps(pie_chart_labels),
        'pie_chart_data': json.dumps(pie_chart_data),
        'radar_labels': json.dumps(radar_labels),
        'radar_data': json.dumps(radar_data),
        'agrupacion': agrupacion,
    }
    return render(request, 'contabilidad_loslirios/analisis.html', context)

#Logic for analisis_movimientos page:
@permission_required('contabilidad_loslirios.can_view_analisis_data', raise_exception=True)
@login_required 
def analisis_movimientos(request):
    # Usamos la función auxiliar para obtener el queryset filtrado
    queryset = _get_movimientos_filtrados_queryset(request)
    form = FormFiltroDashboardMovimientos(request.GET or None)

    # --- CALCULAR KPIs ---
    gasto_total = queryset.aggregate(total=Coalesce(Sum('monto'), Decimal('0.0'))).get('total', Decimal('0.0'))
    gasto_energia = queryset.filter(tipo='Energia').aggregate(total=Coalesce(Sum('monto'), Decimal('0.0'))).get('total', Decimal('0.0'))
    gasto_sueldos_personal = queryset.filter(tipo='Sueldos Personal').aggregate(total=Coalesce(Sum('monto'), Decimal('0.0'))).get('total', Decimal('0.0'))
    gasto_inversion = queryset.filter(tipo='Inversion').aggregate(total=Coalesce(Sum('monto'), Decimal('0.0'))).get('total', Decimal('0.0'))
    gasto_oficial_calculado = queryset.filter(origen='Oficial').aggregate(total=Coalesce(Sum('monto'), Decimal('0.0'))).get('total', Decimal('0.0'))
    gasto_no_oficial_calculado = queryset.filter(origen='No Oficial').aggregate(total=Coalesce(Sum('monto'), Decimal('0.0'))).get('total', Decimal('0.0'))

    # Calcular porcentajes
    porcentaje_oficial = (gasto_oficial_calculado / gasto_total * 100) if gasto_total > 0 else 0
    porcentaje_no_oficial = (gasto_no_oficial_calculado / gasto_total * 100) if gasto_total > 0 else 0

    kpis = {
        'gasto_oficial': gasto_oficial_calculado,
        'gasto_no_oficial': gasto_no_oficial_calculado,
        'gasto_total': gasto_total,
        'porcentaje_oficial': porcentaje_oficial,
        'porcentaje_no_oficial': porcentaje_no_oficial,
        'porcentaje_energia': (gasto_energia / gasto_total * 100) if gasto_total > 0 else 0,
        'porcentaje_sueldos_personal': (gasto_sueldos_personal / gasto_total * 100) if gasto_total > 0 else 0,
        'porcentaje_inversion': (gasto_inversion / gasto_total * 100) if gasto_total > 0 else 0,
        'iva': gasto_oficial_calculado * Decimal('0.21'),
    }

    # --- DATOS PARA GRÁFICOS ---
    # Gráfico de Barras: Top 5 Clasificaciones
    top5_clasificaciones = list(queryset.values('clasificacion').annotate(
        total_monto=Sum('monto')
    ).order_by('-total_monto')[:5])
    
    top5_bar_chart_labels = [g['clasificacion'] for g in top5_clasificaciones]
    top5_bar_chart_data = [float(g['total_monto']) for g in top5_clasificaciones]
    
    # Gráfico de Torta: Distribución de Gastos por Finca
    gastos_por_finca = list(queryset.values('finca').annotate(
        total_monto=Sum('monto')
    ).order_by('-total_monto'))
    
    pie_chart_labels = [g['finca'] for g in gastos_por_finca]
    pie_chart_data = [float(g['total_monto']) for g in gastos_por_finca]

    # Gráfico de Barras: Gastos Totales por Tipo
    gastos_por_tipo = list(queryset.values('tipo').annotate(
        total_monto=Sum('monto')
    ).order_by('-total_monto'))
    
    bar_chart_labels = [g['tipo'] for g in gastos_por_tipo]
    bar_chart_data = [float(g['total_monto']) for g in gastos_por_tipo]

    # Simular datos de alertas (puedes mejorarlo después)
    alertas = {
        'aumentos_significativos': [
            # {'tipo': 'Energía', 'aumento': 15.5},
            # {'tipo': 'Combustibles', 'aumento': 8.2}
        ]
    }
    
    estacionalidad = {
        'mes_pico': 'Enero',
        'monto_pico': gasto_total * Decimal('1.2'),
        'mes_bajo': 'Julio',
        'monto_bajo': gasto_total * Decimal('0.8')
    }

    # Debug: Imprimir datos para verificar
    print("DEBUG - Datos de gráficos:")
    print(f"Top5 labels: {top5_bar_chart_labels}")
    print(f"Top5 data: {top5_bar_chart_data}")
    print(f"Pie labels: {pie_chart_labels}")
    print(f"Pie data: {pie_chart_data}")
    print(f"Bar labels: {bar_chart_labels}")
    print(f"Bar data: {bar_chart_data}")

    # --- CONTEXTO ---
    context = {
        'form': form,
        'kpis': kpis,
        'alertas': alertas,
        'estacionalidad': estacionalidad,
        'top5_bar_chart_labels': json.dumps(top5_bar_chart_labels),
        'top5_bar_chart_data': json.dumps(top5_bar_chart_data),
        'pie_chart_labels': json.dumps(pie_chart_labels),
        'pie_chart_data': json.dumps(pie_chart_data),
        'bar_chart_labels': json.dumps(bar_chart_labels),      
        'bar_chart_data': json.dumps(bar_chart_data),          
    }
    return render(request, 'contabilidad_loslirios/visualizacion/analisis_movimientos.html', context)

#Logic for line_chart_data_api
def _get_movimientos_filtrados_queryset(request):
    """Función auxiliar para obtener el queryset de movimientos filtrado."""
    form = FormFiltroDashboardMovimientos(request.GET or None)
    queryset = MovimientoFinanciero.objects.all()

    if form.is_valid():
        filtros = Q()
        for field, value in form.cleaned_data.items():
            if value:
                if field == 'fecha_desde':
                    filtros &= Q(fecha__gte=value)
                elif field == 'fecha_hasta':
                    filtros &= Q(fecha__lte=value)
                else:
                    filtros &= Q(**{f'{field}__in': value})
        queryset = queryset.filter(filtros)
    return queryset

def line_chart_data_api(request):
    """API que devuelve los datos para el gráfico de líneas, con filtros y agrupación."""
    queryset = _get_movimientos_filtrados_queryset(request)
    
    agrupacion = request.GET.get('agrupacion', 'mes')
    trunc_func = TruncMonth('fecha') # Por defecto
    date_format_func = lambda d: d.strftime('%b %Y')

    if agrupacion == 'anio':
        trunc_func = TruncYear('fecha')
        date_format_func = lambda d: d.strftime('%Y')
    elif agrupacion == 'trimestre':
        trunc_func = TruncQuarter('fecha')
        date_format_func = lambda d: f"T{((d.month-1)//3)+1} {d.year}"
    elif agrupacion == 'dia':
        trunc_func = TruncDay('fecha')
        date_format_func = lambda d: d.strftime('%d/%m/%Y')

    gastos_agrupados = queryset.annotate(periodo=trunc_func).values('periodo').annotate(total_monto=Sum('monto')).order_by('periodo')
    
    labels = [date_format_func(g['periodo']) for g in gastos_agrupados]
    data = [float(g['total_monto']) for g in gastos_agrupados]

    return JsonResponse({'labels': labels, 'data': data})



#Logic for flujo de caja page:
# Logic for Flujo Anual page
@login_required 
def flujo_anual(request):
    year_seleccionado = datetime.now().year

    # Ingresos
    ingresos_query = IngresoFinanciero.objects.filter(fecha__year=year_seleccionado)\
        .annotate(month=TruncMonth('fecha'))\
        .values('detalle', 'month')\
        .annotate(total_mes=Sum('monto'))\
        .order_by('detalle', 'month')

    ingresos_data = defaultdict(lambda: [0]*12)
    ingresos_totales = {}
    for item in ingresos_query:
        mes = item['month'].month
        ingresos_data[item['detalle']][mes-1] = item['total_mes']
    for detalle, montos in ingresos_data.items():
        ingresos_totales[detalle] = sum(montos)

    total_por_mes_ingresos = [sum(ingresos_data[detalle][i] for detalle in ingresos_data) for i in range(12)]
    total_general_ingresos = sum(total_por_mes_ingresos)

    # Egresos
    egresos_query = MovimientoFinanciero.objects.filter(fecha__year=year_seleccionado)\
        .annotate(month=TruncMonth('fecha'))\
        .values('clasificacion', 'month')\
        .annotate(total_mes=Sum('monto'))\
        .order_by('clasificacion', 'month')

    egresos_data = defaultdict(lambda: [0]*12)
    egresos_totales = {}
    for item in egresos_query:
        mes = item['month'].month
        egresos_data[item['clasificacion']][mes-1] = item['total_mes']
    for clasificacion, montos in egresos_data.items():
        egresos_totales[clasificacion] = sum(montos)

    total_por_mes_egresos = [sum(egresos_data[clasificacion][i] for clasificacion in egresos_data) for i in range(12)]
    total_general_egresos = sum(total_por_mes_egresos)

    # Nuevas categorías de egresos
    egresos_categorias = [
        ("Obreros", {"clasificacion": ["Obreros"]}),
        ("Encargados", {"clasificacion": ["Encargados"]}),
        ("Gerenciales", {"clasificacion": ["Gerenciales"]}),
        ("VEP", {"clasificacion": ["VEP"]}),
        ("Insumos Varios", {"tipo": ["Insumos Varios"], "exclude_clasificacion": ["Combustibles"]}),
        ("Combustibles", {"clasificacion": ["Combustibles"]}),
        ("Fertilizantes", {"clasificacion": ["Fertilizantes"]}),
        ("Agroquimicos", {"clasificacion": ["Agroquimicos"]}),
        ("Hidraulica", {"clasificacion": ["Hidraulica"]}),
        ("Movilidad", {"clasificacion": ["Movilidad"]}),
        ("Infraestructura", {"clasificacion": ["Infraestructura"]}),
        ("Energia", {"tipo": ["Energia"]}),
    ]

    egresos_data = {}
    egresos_totales = {}

    for nombre, filtro in egresos_categorias:
        # Caso especial: Insumos Varios por tipo, excluyendo combustibles
        if filtro.get("tipo") and nombre == "Insumos Varios":
            query = MovimientoFinanciero.objects.filter(
                fecha__year=year_seleccionado,
                tipo__in=filtro["tipo"]
            ).exclude(clasificacion__in=filtro.get("exclude_clasificacion", []))\
             .annotate(month=TruncMonth('fecha'))\
             .values('month')\
             .annotate(total_mes=Sum('monto'))\
             .order_by('month')
        # Caso especial: Energia por tipo
        elif filtro.get("tipo"):
            query = MovimientoFinanciero.objects.filter(
                fecha__year=year_seleccionado,
                tipo__in=filtro["tipo"]
            ).annotate(month=TruncMonth('fecha'))\
             .values('month')\
             .annotate(total_mes=Sum('monto'))\
             .order_by('month')
        # Resto de egresos por clasificacion
        else:
            query = MovimientoFinanciero.objects.filter(
                fecha__year=year_seleccionado,
                clasificacion__in=filtro["clasificacion"]
            ).annotate(month=TruncMonth('fecha'))\
             .values('month')\
             .annotate(total_mes=Sum('monto'))\
             .order_by('month')

        montos = [0]*12
        for item in query:
            mes = item['month'].month
            montos[mes-1] = item['total_mes']
        egresos_data[nombre] = montos
        egresos_totales[nombre] = sum(montos)

    total_por_mes_egresos = [sum(egresos_data[nombre][i] for nombre in egresos_data) for i in range(12)]
    total_general_egresos = sum(total_por_mes_egresos)

        # Inversiones 
    inversiones_query = MovimientoFinanciero.objects.filter(
        fecha__year=year_seleccionado,
        tipo='Inversion'
    ).annotate(month=TruncMonth('fecha'))\
     .values('month')\
     .annotate(total_mes=Sum('monto'))\
     .order_by('month')

    inversiones_por_mes = [0]*12
    for item in inversiones_query:
        mes = item['month'].month
        inversiones_por_mes[mes-1] = item['total_mes']
    total_inversiones = sum(inversiones_por_mes)

    # Saldo = Ingresos - Egresos - Inversiones
    saldo_por_mes = [
        total_por_mes_ingresos[i] - total_por_mes_egresos[i] - inversiones_por_mes[i]
        for i in range(12)
    ]
    saldo_total = sum(saldo_por_mes)

    context = {
        'year_seleccionado': year_seleccionado,
        'ingresos_data': dict(ingresos_data),
        'ingresos_totales': ingresos_totales,
        'total_por_mes_ingresos': total_por_mes_ingresos,
        'total_general_ingresos': total_general_ingresos,
        'egresos_data': dict(egresos_data),
        'egresos_totales': egresos_totales,
        'total_por_mes_egresos': total_por_mes_egresos,
        'total_general_egresos': total_general_egresos,
        'inversiones_por_mes': inversiones_por_mes,
        'total_inversiones': total_inversiones,
        'saldo_por_mes': saldo_por_mes,
        'saldo_total': saldo_total,
        'meses': ["Ene", "Feb", "Mar", "Abr", "May", "Jun", "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"],
    }
    return render(request, 'contabilidad_loslirios/flujo/flujo_anual.html', context)

#Logic for Sueldos page:
@login_required 
def sueldos_flujo_anual(request):
    year_seleccionado = datetime.now().year

    meses_nombre_corto = ["Ene", "Feb", "Mar", "Abr", "May", "Jun", "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"]

    # Estructura base (SIN ESPACIOS EN LAS CLAVES)
    sueldos_data = {
        'ENCARGADO_FINCA': {},
        'OBREROS': {
            'TAREA GENERAL': [0] * 12,
            'TAREA OTOÑO': [0] * 12,
            'TAREA INVIERNO': [0] * 12,
            'TAREA PRIMAVERA': [0] * 12,
            'TAREA COSECHA': [0] * 12,
        },
        'GERENCIALES': {},
        'VEP': {'VEP': [0] * 12},
    }

    # 1) Obtener pagos de sueldos desde MovimientoFinanciero
    sueldos_qs = (
        MovimientoFinanciero.objects
        .filter(tipo='Sueldos Personal', fecha__year=year_seleccionado)
        .annotate(month=TruncMonth('fecha'))
        .values('clasificacion', 'detalle', 'month')
        .annotate(total_mes=Sum('monto'))
        .order_by('clasificacion', 'detalle', 'month')
    )

    for item in sueldos_qs:
        clas = (item.get('clasificacion') or '').strip()
        detalle = (item.get('detalle') or 'Sin Nombre').strip()
        month = item['month'].month - 1
        total = item['total_mes'] or 0

        if clas.lower().startswith('encarg'):
            target = 'ENCARGADO_FINCA'
            if detalle not in sueldos_data[target]:
                sueldos_data[target][detalle] = [0] * 12
            sueldos_data[target][detalle][month] += total
        elif clas.lower().startswith('gerenc'):
            target = 'GERENCIALES'
            if detalle not in sueldos_data[target]:
                sueldos_data[target][detalle] = [0] * 12
            sueldos_data[target][detalle][month] += total
        elif clas.lower() == 'vep' or clas.upper() == 'VEP':
            sueldos_data['VEP']['VEP'][month] += total
        elif clas.lower().startswith('obrer'):
            sueldos_data['OBREROS'].setdefault('TAREA GENERAL', [0] * 12)
            sueldos_data['OBREROS']['TAREA GENERAL'][month] += total
        else:
            if detalle not in sueldos_data['GERENCIALES']:
                sueldos_data['GERENCIALES'][detalle] = [0] * 12
            sueldos_data['GERENCIALES'][detalle][month] += total

    # 2) Agregar jornales
    jornales_qs = (
        registro_trabajo.objects
        .filter(fecha__year=year_seleccionado)
        .annotate(month=TruncMonth('fecha'))
        .values('clasificacion', 'month')
        .annotate(total_mes=Sum('monto_total'))
        .order_by('clasificacion', 'month')
    )

    temporada_to_fila = {
        'General': 'TAREA GENERAL',
        'Otoño': 'TAREA OTOÑO',
        'Invierno': 'TAREA INVIERNO',
        'Primavera': 'TAREA PRIMAVERA',
        'Verano': 'TAREA COSECHA',
    }

    for item in jornales_qs:
        temporada = item.get('clasificacion') or 'General'
        fila = temporada_to_fila.get(temporada, 'TAREA GENERAL')
        mes = item['month'].month - 1
        total = item['total_mes'] or 0
        if fila not in sueldos_data['OBREROS']:
            sueldos_data['OBREROS'][fila] = [0] * 12
        sueldos_data['OBREROS'][fila][mes] += total

    # 3) Calcular totales por fila
    sueldos_totales = {}
    for categoria, datos in sueldos_data.items():
        sueldos_totales[categoria] = {}
        for nombre, meses_vals in datos.items():
            sueldos_totales[categoria][nombre] = sum(meses_vals)

    context = {
        'year_seleccionado': year_seleccionado,
        'sueldos_data': sueldos_data,
        'sueldos_totales': sueldos_totales,
        'meses': meses_nombre_corto,
    }
    return render(request, 'contabilidad_loslirios/visualizacion/sueldos_flujo_anual.html', context)



# === EDITAR/ELIMINAR MOVIMIENTOS FINANCIEROS ===
@login_required
@permission_required('contabilidad_loslirios.can_add_movimientos', raise_exception=True)
def editar_movimiento(request, id):
    # CORREGIDO: usar id_movimiento en lugar de id
    movimiento = get_object_or_404(MovimientoFinanciero, id_movimiento=id)
    
    if request.method == 'POST':
        formulario = FormMovimientoFinanciero(request.POST, instance=movimiento)
        if formulario.is_valid():
            formulario.save()
            messages.success(request, f'Movimiento financiero actualizado exitosamente.')
            return redirect('consultar_movimiento')
    else:
        formulario = FormMovimientoFinanciero(instance=movimiento)
    
    context = {
        'formulario': formulario,
        'movimiento': movimiento,
        'modo': 'editar'
    }
    return render(request, 'contabilidad_loslirios/administracion/cargar_movimiento.html', context)

@login_required
@permission_required('contabilidad_loslirios.can_add_movimientos', raise_exception=True)
def eliminar_movimiento(request, id):
    # CORREGIDO: usar id_movimiento en lugar de id
    movimiento = get_object_or_404(MovimientoFinanciero, id_movimiento=id)
    
    if request.method == 'POST':
        descripcion = f"{movimiento.fecha.strftime('%d/%m/%Y')} - {movimiento.tipo} - ${movimiento.monto:.2f}"
        movimiento.delete()
        messages.success(request, f'Movimiento "{descripcion}" eliminado exitosamente.')
        return redirect('consultar_movimiento')
    
    context = {'movimiento': movimiento}
    return render(request, 'contabilidad_loslirios/administracion/confirmar_eliminar_movimiento.html', context)

# === EDITAR/ELIMINAR INGRESOS FINANCIEROS ===
@login_required
@permission_required('contabilidad_loslirios.can_add_ingresos', raise_exception=True)
def editar_ingreso(request, id):
    # CORREGIDO: usar id_ingreso en lugar de id
    ingreso = get_object_or_404(IngresoFinanciero, id_ingreso=id)
    
    if request.method == 'POST':
        formulario = FormIngresoFinanciero(request.POST, instance=ingreso)
        if formulario.is_valid():
            formulario.save()
            messages.success(request, f'Ingreso actualizado exitosamente.')
            return redirect('consultar_ingresos')
    else:
        formulario = FormIngresoFinanciero(instance=ingreso)
    
    context = {
        'formulario': formulario,
        'ingreso': ingreso,
        'modo': 'editar'
    }
    return render(request, 'contabilidad_loslirios/administracion/cargar_ingresos.html', context)

@login_required
@permission_required('contabilidad_loslirios.can_add_ingresos', raise_exception=True)
def eliminar_ingreso(request, id):
    # CORREGIDO: usar id_ingreso en lugar de id
    ingreso = get_object_or_404(IngresoFinanciero, id_ingreso=id)
    
    if request.method == 'POST':
        descripcion = f"{ingreso.comprador} - {ingreso.fecha.strftime('%d/%m/%Y')} - ${ingreso.monto:.2f}"
        ingreso.delete()
        messages.success(request, f'Ingreso "{descripcion}" eliminado exitosamente.')
        return redirect('consultar_ingresos')
    
    context = {'ingreso': ingreso}
    return render(request, 'contabilidad_loslirios/administracion/confirmar_eliminar_ingreso.html', context)

# === EDITAR/ELIMINAR JORNALES ===
@login_required
@permission_required('contabilidad_loslirios.can_add_jornales', raise_exception=True)
def editar_jornal(request, id):
    # CORREGIDO: usar id_registro en lugar de id
    jornal = get_object_or_404(registro_trabajo, id_registro=id)
    
    if request.method == 'POST':
        formulario = FormRegistroTrabajo(request.POST, instance=jornal)
        if formulario.is_valid():
            formulario.save()
            messages.success(request, f'Jornal actualizado exitosamente.')
            return redirect('consultar_jornal')
    else:
        formulario = FormRegistroTrabajo(instance=jornal)
    
    context = {
        'formulario': formulario,
        'jornal': jornal,
        'modo': 'editar'
    }
    return render(request, 'contabilidad_loslirios/produccion/cargar_jornal.html', context)

@login_required
@permission_required('contabilidad_loslirios.can_add_jornales', raise_exception=True)
def eliminar_jornal(request, id):
    # CORREGIDO: usar id_registro en lugar de id
    jornal = get_object_or_404(registro_trabajo, id_registro=id)
    
    if request.method == 'POST':
        descripcion = f"{jornal.nombre_trabajador} - {jornal.fecha.strftime('%d/%m/%Y')} - {jornal.tarea}"
        jornal.delete()
        messages.success(request, f'Jornal "{descripcion}" eliminado exitosamente.')
        return redirect('consultar_jornal')
    
    context = {'jornal': jornal}
    return render(request, 'contabilidad_loslirios/produccion/confirmar_eliminar_jornal.html', context)

# === EDITAR/ELIMINAR REGISTROS DE RIEGO ===
@login_required
@permission_required('contabilidad_loslirios.can_add_riego', raise_exception=True)
def editar_riego(request, id):
    # CORREGIDO: usar id_riego en lugar de id
    riego = get_object_or_404(RegistroRiego, id_riego=id)
    
    if request.method == 'POST':
        formulario = FormRegistroRiego(request.POST, instance=riego)
        if formulario.is_valid():
            formulario.save()
            messages.success(request, f'Registro de riego actualizado exitosamente.')
            return redirect('consultar_riego')
    else:
        formulario = FormRegistroRiego(instance=riego)
    
    context = {
        'form': formulario,
        'riego': riego,
        'modo': 'editar'
    }
    return render(request, 'contabilidad_loslirios/produccion/cargar_riego.html', context)

@login_required
@permission_required('contabilidad_loslirios.can_add_riego', raise_exception=True)
def eliminar_riego(request, id):
    # CORREGIDO: usar id_riego en lugar de id
    riego = get_object_or_404(RegistroRiego, id_riego=id)
    
    if request.method == 'POST':
        descripcion = f"{riego.parral} - {riego.inicio.strftime('%d/%m/%Y %H:%M')}"
        riego.delete()
        messages.success(request, f'Registro de riego "{descripcion}" eliminado exitosamente.')
        return redirect('consultar_riego')
    
    context = {'riego': riego}
    return render(request, 'contabilidad_loslirios/produccion/confirmar_eliminar_riego.html', context)

# === EDITAR/ELIMINAR REGISTROS DE COSECHA ===
@login_required
@permission_required('contabilidad_loslirios.add_registrocosecha', raise_exception=True)
def editar_cosecha(request, id):
    cosecha = get_object_or_404(RegistroCosecha, id=id)
    
    if request.method == 'POST':
        formulario = RegistroCosechaForm(request.POST, instance=cosecha)
        if formulario.is_valid():
            formulario.save()
            messages.success(request, f'Registro de cosecha actualizado exitosamente.')
            return redirect('consultar_cosecha')
    else:
        formulario = RegistroCosechaForm(instance=cosecha)
    
    # Obtener opciones para datalists (igual que en cargar_cosecha)
    fincas = list(set([choice[0] for choice in FINCA_CHOICES] + 
                     list(RegistroCosecha.objects.values_list('finca', flat=True).distinct())))
    compradores = list(set([choice[0] for choice in COMPRADOR_CHOICES] + 
                          list(RegistroCosecha.objects.values_list('comprador', flat=True).distinct())))
    cultivos = list(set([choice[0] for choice in CULTIVO_CHOICES] + 
                       list(RegistroCosecha.objects.values_list('cultivo', flat=True).distinct())))
    variedades = list(set([choice[0] for choice in VARIEDAD_CHOICES] + 
                         list(RegistroCosecha.objects.values_list('variedad', flat=True).distinct())))
    
    context = {
        'form': formulario,
        'cosecha': cosecha,
        'modo': 'editar',
        'fincas': fincas,
        'compradores': compradores,
        'cultivos': cultivos,
        'variedades': variedades,
    }
    return render(request, 'contabilidad_loslirios/produccion/cargar_cosecha.html', context)

@login_required
@permission_required('contabilidad_loslirios.add_registrocosecha', raise_exception=True)
def eliminar_cosecha(request, id):
    cosecha = get_object_or_404(RegistroCosecha, id=id)
    
    if request.method == 'POST':
        descripcion = f"{cosecha.variedad} - {cosecha.fecha.strftime('%d/%m/%Y')}"
        cosecha.delete()
        messages.success(request, f'Registro de cosecha "{descripcion}" eliminado exitosamente.')
        return redirect('consultar_cosecha')
    
    context = {'cosecha': cosecha}
    return render(request, 'contabilidad_loslirios/produccion/confirmar_eliminar_cosecha.html', context)

# === BÚSQUEDA AVANZADA ===
@login_required
@permission_required('contabilidad_loslirios.can_view_movimientos', raise_exception=True)
def busqueda_avanzada_movimientos(request):
    movimientos = MovimientoFinanciero.objects.all()
    
    # Filtros avanzados
    fecha_desde = request.GET.get('fecha_desde')
    fecha_hasta = request.GET.get('fecha_hasta')
    descripcion = request.GET.get('descripcion')
    monto_min = request.GET.get('monto_min')
    monto_max = request.GET.get('monto_max')
    tipo = request.GET.get('tipo')
    
    if fecha_desde:
        movimientos = movimientos.filter(fecha__gte=fecha_desde)
    if fecha_hasta:
        movimientos = movimientos.filter(fecha__lte=fecha_hasta)
    if descripcion:
        movimientos = movimientos.filter(descripcion__icontains=descripcion)
    if monto_min:
        movimientos = movimientos.filter(monto__gte=monto_min)
    if monto_max:
        movimientos = movimientos.filter(monto__lte=monto_max)
    if tipo:
        movimientos = movimientos.filter(tipo=tipo)
    
    # Ordenamiento
    orden = request.GET.get('orden', '-fecha')
    movimientos = movimientos.order_by(orden)
    
    # Paginación
    from django.core.paginator import Paginator
    paginator = Paginator(movimientos, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'movimientos': page_obj,
        'filtros': {
            'fecha_desde': fecha_desde,
            'fecha_hasta': fecha_hasta,
            'descripcion': descripcion,
            'monto_min': monto_min,
            'monto_max': monto_max,
            'tipo': tipo,
        },
        'total_registros': movimientos.count(),
        'total_monto': movimientos.aggregate(total=Sum('monto'))['total'] or 0,
    }
    
    return render(request, 'contabilidad_loslirios/administracion/busqueda_avanzada_movimientos.html', context)

# === EXPORTAR A EXCEL ===
@login_required
@permission_required('contabilidad_loslirios.can_export_movimientos', raise_exception=True)
def exportar_excel_movimientos(request):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from django.http import HttpResponse
    from datetime import datetime
    
    # Crear workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Movimientos Financieros"
    
    # Aplicar mismos filtros que en la búsqueda
    movimientos = MovimientoFinanciero.objects.all()
    
    fecha_desde = request.GET.get('fecha_desde')
    fecha_hasta = request.GET.get('fecha_hasta')
    descripcion = request.GET.get('descripcion')
    tipo = request.GET.get('tipo')
    
    if fecha_desde:
        movimientos = movimientos.filter(fecha__gte=fecha_desde)
    if fecha_hasta:
        movimientos = movimientos.filter(fecha__lte=fecha_hasta)
    if descripcion:
        movimientos = movimientos.filter(descripcion__icontains=descripcion)
    if tipo:
        movimientos = movimientos.filter(tipo=tipo)
    
    movimientos = movimientos.order_by('-fecha')
    
    # Configurar encabezados
    headers = ['Fecha', 'Descripción', 'Tipo', 'Monto', 'Forma de Pago', 'Número Cheque', 'Vencimiento Cheque']
    
    # Estilo para encabezados
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Escribir encabezados
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Escribir datos
    for row, movimiento in enumerate(movimientos, 2):
        ws.cell(row=row, column=1, value=movimiento.fecha.strftime('%d/%m/%Y'))
        ws.cell(row=row, column=2, value=movimiento.descripcion)
        ws.cell(row=row, column=3, value=movimiento.tipo)
        ws.cell(row=row, column=4, value=float(movimiento.monto))
        ws.cell(row=row, column=5, value=movimiento.forma_pago)
        ws.cell(row=row, column=6, value=movimiento.numero_cheque or '')
        
        if movimiento.vencimiento_cheque:
            ws.cell(row=row, column=7, value=movimiento.vencimiento_cheque.strftime('%d/%m/%Y'))
    
    # Ajustar ancho de columnas
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Agregar totales al final
    total_row = len(movimientos) + 3
    ws.cell(row=total_row, column=3, value='Total')
    ws.cell(row=total_row, column=4, value=f'=SUM(D2:D{total_row-1})').font = Font(bold=True)
    
    # Proteger la hoja
    ws.protect("tu_contraseña", sheet=True)
    
    # Desbloquear celdas específicas
    for row in range(2, total_row):
        for col in [1, 2, 3, 5, 6, 7]:  # Desbloquear columnas A, B, C, E, F, G
            ws.cell(row=row, column=col).protection = None
    
    # Forzar descarga del archivo
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="movimientos_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    
    # Guardar el archivo en la respuesta
    wb.save(response)
    return response

# Logic for Dashboard KPIs API
@login_required
def dashboard_kpis_api(request):
    """API para obtener KPIs del dashboard en tiempo real"""
    from datetime import datetime, timedelta
    from django.db.models import Sum
    import traceback
    
    try:
        hoy = datetime.now().date()
        primer_dia_mes = hoy.replace(day=1)
        ultimo_dia_mes_anterior = primer_dia_mes - timedelta(days=1)
        primer_dia_mes_anterior = ultimo_dia_mes_anterior.replace(day=1)
        
        print(f"DEBUG - Consultando datos desde {primer_dia_mes} hasta {hoy}")  # DEBUG
        
        # GASTOS DEL MES ACTUAL (MOVIMIENTOS FINANCIEROS)
        gastos_mes = MovimientoFinanciero.objects.filter(
            fecha__gte=primer_dia_mes,
            fecha__lte=hoy
        ).aggregate(total=Sum('monto'))['total'] or 0
        
        # INGRESOS DEL MES ACTUAL (INGRESOS FINANCIEROS)
        ingresos_mes = IngresoFinanciero.objects.filter(
            fecha__gte=primer_dia_mes,
            fecha__lte=hoy
        ).aggregate(total=Sum('monto'))['total'] or 0
        
        # CHEQUES DEL MES ACTUAL
        cheques_mes_queryset = IngresoFinanciero.objects.filter(
            fecha__gte=primer_dia_mes,
            fecha__lte=hoy,
            forma_pago__in=['Cheque', 'Echeque']
        )
        
        # TRABAJADORES DEL MES ACTUAL
        trabajadores_mes = registro_trabajo.objects.filter(
            fecha__gte=primer_dia_mes,
            fecha__lte=hoy
        ).values('nombre_trabajador').distinct().count()
        
        # TRABAJADORES DE LA SEMANA ACTUAL
        inicio_semana = hoy - timedelta(days=hoy.weekday())
        trabajadores_semana = registro_trabajo.objects.filter(
            fecha__gte=inicio_semana,
            fecha__lte=hoy
        ).values('nombre_trabajador').distinct().count()
        
        # GASTOS DEL MES ANTERIOR PARA CALCULAR VARIACIÓN
        gastos_mes_anterior = MovimientoFinanciero.objects.filter(
            fecha__gte=primer_dia_mes_anterior,
            fecha__lte=ultimo_dia_mes_anterior
        ).aggregate(total=Sum('monto'))['total'] or 0
        
        # INGRESOS DEL MES ANTERIOR PARA CALCULAR VARIACIÓN
        ingresos_mes_anterior = IngresoFinanciero.objects.filter(
            fecha__gte=primer_dia_mes_anterior,
            fecha__lte=ultimo_dia_mes_anterior
        ).aggregate(total=Sum('monto'))['total'] or 0
        
        # CALCULAR VARIACIONES
        variacion_gastos = 0
        if gastos_mes_anterior > 0:
            variacion_gastos = ((gastos_mes - gastos_mes_anterior) / gastos_mes_anterior * 100)
        
        variacion_ingresos = 0
        if ingresos_mes_anterior > 0:
            variacion_ingresos = ((ingresos_mes - ingresos_mes_anterior) / ingresos_mes_anterior * 100)
        
        # SPARKLINES (ÚLTIMOS 5 DÍAS) - CORREGIDO
        sparkline_gastos = []
        sparkline_ingresos = []
        
        if gastos_mes > 0 or ingresos_mes > 0:
            for i in range(4, -1, -1):
                fecha = hoy - timedelta(days=i)
                gasto_dia = MovimientoFinanciero.objects.filter(fecha=fecha).aggregate(total=Sum('monto'))['total'] or 0
                ingreso_dia = IngresoFinanciero.objects.filter(fecha=fecha).aggregate(total=Sum('monto'))['total'] or 0
                sparkline_gastos.append(float(gasto_dia))
                sparkline_ingresos.append(float(ingreso_dia))
        
        # DEBUG: Mostrar datos calculados
        print(f"DEBUG - Gastos mes actual: {gastos_mes}")
        print(f"DEBUG - Ingresos mes actual: {ingresos_mes}")
        print(f"DEBUG - Trabajadores mes: {trabajadores_mes}")
        print(f"DEBUG - Cheques mes: {cheques_mes_queryset.count()}")
        
        # PREPARAR RESPUESTA
        data = {
            'gasto_total_mes': float(gastos_mes),
            'ingresos_mes': float(ingresos_mes),
            'saldo_disponible': float(ingresos_mes - gastos_mes),
            'cheques_mes': cheques_mes_queryset.count(),
            'monto_cheques_mes': float(cheques_mes_queryset.aggregate(total=Sum('monto'))['total'] or 0),
            'trabajadores_mes': trabajadores_mes,
            'trabajadores_semana': trabajadores_semana,
            'variacion_gastos': float(variacion_gastos),
            'variacion_ingresos': float(variacion_ingresos),
            'sparkline_gastos': sparkline_gastos,
            'sparkline_ingresos': sparkline_ingresos,
            'tiene_datos': gastos_mes > 0 or ingresos_mes > 0 or trabajadores_mes > 0,
        }
        
        print(f"DEBUG - Datos finales enviados: {data}")
        
        return JsonResponse(data)
        
    except Exception as e:
        print(f"ERROR en dashboard_kpis_api: {str(e)}")
        print(traceback.format_exc())
        return JsonResponse({'error': str(e)}, status=500)
